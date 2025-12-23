#!/usr/bin/env python3
"""
Import recipes from Excel file (source.xlsx) into storage_data.json
Each sheet represents a meal with ingredients and per-person portions
"""
import json
from pathlib import Path
from openpyxl import load_workbook

DATA_DIR = Path(__file__).parent.parent / "data"
EXCEL_FILE = DATA_DIR / "source.xlsx"
JSON_FILE = DATA_DIR / "storage_data.json"


def normalize_name(name: str) -> str:
    """Normalize ingredient name for matching."""
    return name.lower().strip()


def load_json_data() -> dict:
    """Load existing storage data."""
    with open(JSON_FILE, 'r') as f:
        return json.load(f)


def save_json_data(data: dict):
    """Save storage data to JSON."""
    with open(JSON_FILE, 'w') as f:
        json.dump(data, f, indent=2)


def get_next_recipe_id(data: dict) -> int:
    """Get the next available recipe ID."""
    if not data['recipes']:
        return 1
    return max(recipe['id'] for recipe in data['recipes']) + 1


def find_ingredient_by_name(ingredients: list, name: str) -> dict:
    """Find ingredient by name (case-insensitive)."""
    norm_name = normalize_name(name)
    for ing in ingredients:
        if normalize_name(ing['name']) == norm_name:
            return ing
    return None


def read_recipes_from_excel(data: dict):
    """Read recipes from all sheets in Excel file."""
    wb = load_workbook(EXCEL_FILE, read_only=True)
    recipes = []
    skipped_sheets = []

    print(f"Processing {len(wb.sheetnames)} sheets...\n")

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"Reading sheet: {sheet_name}")

        # Row 2, Column B contains the recipe name
        recipe_name_cell = sheet.cell(row=2, column=2)
        recipe_name = recipe_name_cell.value if recipe_name_cell.value else sheet_name

        # Skip if no recipe name
        if not recipe_name or recipe_name == sheet_name:
            print(f"  ⊘ Skipping - no recipe name found")
            skipped_sheets.append(sheet_name)
            continue

        print(f"  Recipe: {recipe_name}")

        # Read ingredients starting from row 7
        recipe_ingredients = []
        missing_ingredients = []
        row_num = 7

        for row in sheet.iter_rows(min_row=7, values_only=True):
            # Column A (index 0) = ingredient name
            # Column D (index 3) = content each meal meat (per person)
            # Column E (index 4) = content each meal vegi (per person)

            if not row or not row[0]:  # Skip empty rows
                continue

            ingredient_name = str(row[0]).strip()

            # Skip invalid entries
            if ingredient_name.lower() in ['none', '', 'as usual'] or ingredient_name == 'None':
                continue

            # Get per-person portions (in kg from Excel)
            portion_meat = row[3] if len(row) > 3 and row[3] else 0
            portion_vegi = row[4] if len(row) > 4 and row[4] else 0

            # Use meat portion as default, or average if both exist
            if portion_meat and portion_vegi:
                portion_kg = (float(portion_meat) + float(portion_vegi)) / 2
            elif portion_meat:
                portion_kg = float(portion_meat)
            elif portion_vegi:
                portion_kg = float(portion_vegi)
            else:
                # No portion specified, skip
                continue

            # Convert kg to grams
            portion_grams = portion_kg * 1000

            # Find matching ingredient in database
            matching_ingredient = find_ingredient_by_name(data['ingredients'], ingredient_name)

            if matching_ingredient:
                recipe_ingredients.append({
                    'ingredient_id': matching_ingredient['id'],
                    'quantity_grams': round(portion_grams, 1)
                })
                print(f"    ✓ {ingredient_name}: {portion_grams:.1f}g per person")
            else:
                missing_ingredients.append(ingredient_name)
                print(f"    ⚠ {ingredient_name}: NOT FOUND in ingredient database")

            row_num += 1

        # Only add recipe if it has at least one ingredient
        if recipe_ingredients:
            # Add comments about missing ingredients
            comments = ""
            if missing_ingredients:
                comments = f"Missing ingredients not in database: {', '.join(missing_ingredients)}"

            recipes.append({
                'name': recipe_name,
                'comments': comments,
                'ingredients': recipe_ingredients,
                'sheet': sheet_name
            })
            print(f"  ✓ Added recipe with {len(recipe_ingredients)} ingredients")
        else:
            print(f"  ⊘ Skipping - no valid ingredients found")
            skipped_sheets.append(sheet_name)

        print()

    return recipes, skipped_sheets


def main():
    print("=" * 60)
    print("Starting recipe import from Excel...")
    print("=" * 60)

    # Load existing data
    data = load_json_data()
    print(f"\nCurrent recipes in storage: {len(data['recipes'])}")
    print(f"Available ingredients: {len(data['ingredients'])}\n")

    # Read recipes from Excel
    recipes, skipped_sheets = read_recipes_from_excel(data)

    if not recipes:
        print("No recipes found to import!")
        return

    # Check for duplicate recipe names
    existing_recipe_names = {r['name'].lower() for r in data['recipes']}
    new_recipes = []
    skipped_duplicates = []

    next_id = get_next_recipe_id(data)

    for recipe_data in recipes:
        recipe_name = recipe_data['name']

        # Skip duplicates
        if recipe_name.lower() in existing_recipe_names:
            print(f"⊘ Skipping duplicate recipe: '{recipe_name}'")
            skipped_duplicates.append(recipe_name)
            continue

        # Add recipe with ID
        new_recipe = {
            'id': next_id,
            'name': recipe_name,
            'comments': recipe_data['comments'],
            'ingredients': recipe_data['ingredients']
        }

        data['recipes'].append(new_recipe)
        new_recipes.append(recipe_name)
        print(f"✓ Added recipe: '{recipe_name}' (ID {next_id}, {len(recipe_data['ingredients'])} ingredients)")
        next_id += 1

    # Save updated data
    if new_recipes:
        save_json_data(data)

    print("\n" + "=" * 60)
    print("=== IMPORT SUMMARY ===")
    print(f"Sheets processed: {len(wb.sheetnames)}")
    print(f"Recipes added: {len(new_recipes)}")
    print(f"Skipped (duplicates): {len(skipped_duplicates)}")
    print(f"Skipped (no data): {len(skipped_sheets)}")
    print(f"Total recipes in storage: {len(data['recipes'])}")
    print("=" * 60)

    if new_recipes:
        print("\nAdded recipes:")
        for recipe_name in new_recipes:
            print(f"  - {recipe_name}")

    if skipped_duplicates:
        print("\nSkipped duplicates:")
        for recipe_name in skipped_duplicates:
            print(f"  - {recipe_name}")


if __name__ == "__main__":
    main()
