#!/usr/bin/env python3
"""
Import ingredients from Excel file (source.xlsx) into storage_data.json
"""
import json
from pathlib import Path
from openpyxl import load_workbook

DATA_DIR = Path(__file__).parent.parent / "data"
EXCEL_FILE = DATA_DIR / "source.xlsx"
JSON_FILE = DATA_DIR / "storage_data.json"

# Category mapping based on keywords
CATEGORY_KEYWORDS = {
    "Vegetables": ["tomato", "potato", "onion", "carrot", "lettuce", "spinach", "cucumber",
                  "pepper", "cabbage", "broccoli", "zucchini", "eggplant", "celery", "radish",
                  "garlic", "leek", "parsley", "cilantro", "kale", "chard", "beet"],
    "Fruits": ["apple", "banana", "orange", "grape", "berry", "melon", "peach", "pear",
              "mango", "strawberry", "blueberry", "raspberry", "lemon", "lime", "pineapple"],
    "Meat": ["beef", "chicken", "pork", "lamb", "turkey", "fish", "salmon", "tuna",
            "steak", "bacon", "sausage", "ham", "meat", "fleisch"],
    "Dairy": ["milk", "cheese", "yogurt", "butter", "cream", "eggs", "egg", "ei",
             "käse", "milch", "sahne"],
    "Grains": ["rice", "pasta", "bread", "flour", "oats", "quinoa", "wheat", "noodle",
              "spaghetti", "couscous", "bulgur", "reis", "mehl"],
    "Spices": ["salt", "pepper", "curry", "paprika", "cumin", "oregano", "basil",
              "thyme", "rosemary", "sage", "cinnamon", "turmeric", "salz", "pfeffer"],
    "Beverages": ["water", "juice", "coffee", "tea", "soda", "wine", "beer", "wasser",
                 "saft", "wein"],
    "Canned Goods": ["canned", "tinned", "jar", "dose"],
    "Frozen": ["frozen", "tiefkühl", "tk"],
    "Other": []
}

# Unit mapping to simplified measurements
UNIT_MAPPING = {
    'kg': 'kg', 'kilogram': 'kg', 'g': 'kg', 'gram': 'kg', 'grams': 'kg',
    'l': 'liter', 'liter': 'liter', 'litre': 'liter', 'ml': 'liter', 'milliliter': 'liter',
    'pieces': 'pieces', 'piece': 'pieces', 'pcs': 'pieces', 'count': 'pieces',
    'items': 'pieces', 'cans': 'pieces', 'packages': 'pieces', 'bottles': 'pieces',
    'stk': 'pieces', 'stück': 'pieces', 'stk.': 'pieces'  # German units
}


def map_unit_to_measurement(unit: str) -> str:
    """Map Excel unit to simplified measurement type."""
    if not unit or unit == 'None':
        return 'pieces'
    unit_lower = str(unit).lower().strip()
    return UNIT_MAPPING.get(unit_lower, 'pieces')  # Default to pieces


def smart_map_category(name: str) -> str:
    """Smart category mapping based on ingredient name."""
    name_lower = name.lower()
    for category, keywords in CATEGORY_KEYWORDS.items():
        if category == "Other":
            continue
        if any(keyword in name_lower for keyword in keywords):
            return category
    return "Other"  # Default category


def load_json_data() -> dict:
    """Load existing storage data."""
    if JSON_FILE.exists():
        with open(JSON_FILE, 'r') as f:
            return json.load(f)
    return {
        "ingredients": [],
        "recipes": [],
        "categories": [
            "Vegetables", "Fruits", "Meat", "Dairy", "Grains",
            "Spices", "Beverages", "Canned Goods", "Frozen", "Other"
        ],
        "units": ["kg", "liter", "pieces"]
    }


def save_json_data(data: dict):
    """Save storage data to JSON."""
    with open(JSON_FILE, 'w') as f:
        json.dump(data, f, indent=2)


def get_next_ingredient_id(data: dict) -> int:
    """Get the next available ingredient ID."""
    if not data['ingredients']:
        return 1
    return max(ing['id'] for ing in data['ingredients']) + 1


def read_excel_ingredients():
    """Read ingredients from all sheets in Excel file."""
    wb = load_workbook(EXCEL_FILE, read_only=True)

    ingredients = []
    seen_names = set()  # Track unique ingredients across all sheets

    # Process all sheets
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"Reading sheet: {sheet_name}")

        # Read data rows (starting from row 7)
        row_count = 0
        for row in sheet.iter_rows(min_row=7, values_only=True):
            # Column A (index 0) = ingredient name
            # Column C (index 2) = unit
            if not row or not row[0]:  # Skip empty rows or rows without ingredient name
                continue

            name = str(row[0]).strip()

            # Skip invalid entries
            if name.lower() in ['none', '', 'as usual'] or name == 'None':
                continue

            # Skip duplicates (case-insensitive)
            if name.lower() in seen_names:
                continue

            seen_names.add(name.lower())
            row_count += 1

            # Column C = unit (handle None)
            unit = row[2] if len(row) > 2 and row[2] else 'pieces'

            ingredients.append({
                'name': name,
                'unit': unit,
                'category': None  # Will use smart mapping
            })

        print(f"  Found {row_count} unique ingredients in this sheet")

    print(f"\nExtracted {len(ingredients)} unique ingredients from {len(wb.sheetnames)} sheets")
    return ingredients


def main():
    print("=" * 60)
    print("Starting ingredient import from Excel...")
    print("=" * 60)

    # Load existing data
    data = load_json_data()
    existing_names = {ing['name'].lower() for ing in data['ingredients']}
    print(f"\nCurrent ingredients in storage: {len(data['ingredients'])}")

    # Read Excel
    excel_ingredients = read_excel_ingredients()

    # Import new ingredients
    added = 0
    skipped = 0
    next_id = get_next_ingredient_id(data)

    print(f"\n{'=' * 60}")
    print("Processing ingredients...")
    print("=" * 60)

    for excel_ing in excel_ingredients:
        name = excel_ing['name']

        # Skip duplicates
        if name.lower() in existing_names:
            print(f"⊘ Skipping '{name}' (already exists)")
            skipped += 1
            continue

        # Map measurement
        measurement = map_unit_to_measurement(excel_ing['unit'])

        # Smart category mapping
        category = excel_ing['category'] if excel_ing['category'] else smart_map_category(name)

        # Add ingredient
        ingredient = {
            'id': next_id,
            'name': name,
            'category': category,
            'measurement': measurement,
            'amount': 1.0
        }

        data['ingredients'].append(ingredient)
        print(f"✓ Added: {name} ({category}, {measurement}, 1.0)")
        added += 1
        next_id += 1

    # Save updated data
    if added > 0:
        save_json_data(data)
        print(f"\n{'=' * 60}")
        print("Saved updated data to storage_data.json")
    else:
        print(f"\n{'=' * 60}")
        print("No new ingredients to add - no changes made")

    print("=" * 60)
    print("\n=== IMPORT SUMMARY ===")
    print(f"Total unique ingredients in Excel: {len(excel_ingredients)}")
    print(f"Added: {added}")
    print(f"Skipped (duplicates): {skipped}")
    print(f"Total ingredients in storage: {len(data['ingredients'])}")
    print("=" * 60)


if __name__ == "__main__":
    main()
